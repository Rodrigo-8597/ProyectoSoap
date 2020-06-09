import logging
from itertools import cycle

logging.basicConfig(level=logging.DEBUG)
from spyne import Application, rpc, ServiceBase, \
    Integer, Unicode
from spyne import Iterable
from spyne.protocol.http import HttpRpc
from spyne.protocol.json import JsonDocument
from spyne.server.wsgi import WsgiApplication
from spyne.protocol.soap import Soap11
from spyne.model.primitive import String

from openpyxl import workbook
import base64

class psuService(ServiceBase):
    @rpc(Unicode,_returns = Iterable(Unicode))
    def separacion(ctx, dato_64):
        carreras=[]
        for i in range(0,28):
            carreras.append([])
        
        base64_bytes=dato_64.encode('ascii')
        message_bytes=base64.b64decode(base64_bytes)
        message=message_bytes.decode('ascii')
        message=message.split('\n')
        for i in range(0,len(message)):
            message[i]=message[i].split(';')
        


        wb=Workbook()
        ws=wb.active
        ws.title="C. 1"
        ws1 = wb.create_sheet("C. 2")
        ws2 = wb.create_sheet("C. 3")
        ws3 = wb.create_sheet("C. 4")
        ws4 = wb.create_sheet("C. 5")
        ws5 = wb.create_sheet("C. 6")
        ws6 = wb.create_sheet("C. 7")
        ws7 = wb.create_sheet("C. 8")
        ws8 = wb.create_sheet("C. 9")
        ws9 = wb.create_sheet("C. 10")
        ws10 = wb.create_sheet("C. 11")
        ws11 = wb.create_sheet("C. 12")
        ws12 = wb.create_sheet("C. 13")
        ws13 = wb.create_sheet("C. 14")
        ws14 = wb.create_sheet("C. 15")
        ws15 = wb.create_sheet("C. 16")
        ws16 = wb.create_sheet("C. 17")
        ws17 = wb.create_sheet("C. 18")
        ws18 = wb.create_sheet("C. 19")
        ws19 = wb.create_sheet("C. 20")
        ws20 = wb.create_sheet("C. 21")
        ws21 = wb.create_sheet("C. 22")
        ws22 = wb.create_sheet("C. 23")
        ws23 = wb.create_sheet("C. 24")
        ws24 = wb.create_sheet("C. 25")
        ws25 = wb.create_sheet("C. 26")
        ws26 = wb.create_sheet("C. 27")
        ws27 = wb.create_sheet("C. 28")

        wb.save(filename = 'admitidos.xlsx')
        yield(excel_base64)

class digitoService(ServiceBase):
    @rpc(Unicode, Unicode, _returns = Iterable(Unicode))
    def digito_verificador(ctx, rut, times):
        n_rut = rut.split('-')
        reversed_digits = map(int, reversed(str(n_rut[0])))
        factors = cycle(range(2, 8))
        s = sum(d * f for d, f in zip(reversed_digits, factors))
        mod = (-s) % 11
        if (mod == 10):
            mod = 'k'
        if (mod == 11):
            mod = 0
        if (str(mod) == n_rut[1]):
            yield ('Para el rut ' + str(rut) + ' ' + 'el digito verificador es '+ str(mod))
        else:
            yield('dv ingresado '+ str(n_rut[1]) + ' el dv correcto es '+ str(mod))


class nompropService(ServiceBase):
    @rpc(Unicode, Unicode, Unicode, Unicode, _returns = Iterable(Unicode))
    def generar_saludo(ctx, nom, pat, mat, sexo):
        nombreCompleto = nom + ' ' + pat + ' ' + mat + ' '
        nomComProp = nombreCompleto.title()
        if (int(sexo) == 1):
            sex = 'Sra. '
        else:
            sex = 'Sr. '
        yield (sex + ' ' + nomComProp )




application = Application(
    [
        digitoService,
        nompropService
    ],
    tns = 'spyne.examples.hello.soap',
    in_protocol = Soap11(),
    out_protocol = Soap11()
)

if __name__ == '__main__':
    # You can use any Wsgi server. Here, we chose
    # Python's built-in wsgi server but you're not
    # supposed to use it in production.
    from wsgiref.simple_server import make_server
    wsgi_app = WsgiApplication(application)
    server = make_server('127.0.0.1', 8000, wsgi_app)
    server.serve_forever()
